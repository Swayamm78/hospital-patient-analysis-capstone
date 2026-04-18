import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
import os

# ─── Load Dataset ───────────────────────────────────────────────
df = pd.read_csv('/home/claude/patients.csv')
print("=" * 55)
print("  HOSPITAL PATIENT DATA ANALYSIS")
print("=" * 55)

# ─── 1. Display All Records ──────────────────────────────────────
print("\n📋 PATIENT RECORDS:")
print(df.to_string(index=False))

# ─── 2. Disease Frequency ────────────────────────────────────────
print("\n🦠 DISEASE FREQUENCY:")
disease_count = df['disease'].value_counts().reset_index()
disease_count.columns = ['Disease', 'Patient Count']
print(disease_count.to_string(index=False))

# ─── 3. Treatment Cost Statistics ───────────────────────────────
print("\n💰 TREATMENT COST STATISTICS:")
stats = df['treatment_cost'].describe()
print(f"  Average  : ₹{stats['mean']:,.0f}")
print(f"  Min      : ₹{stats['min']:,.0f}")
print(f"  Max      : ₹{stats['max']:,.0f}")
print(f"  Std Dev  : ₹{stats['std']:,.0f}")
print(f"  Median   : ₹{df['treatment_cost'].median():,.0f}")

# ─── 4. Group by Doctor ──────────────────────────────────────────
print("\n👨‍⚕️ PATIENTS PER DOCTOR:")
doctor_group = df.groupby('doctor').agg(
    Patient_Count=('patient_id', 'count'),
    Avg_Cost=('treatment_cost', 'mean')
).reset_index().sort_values('Patient_Count', ascending=False)
doctor_group['Avg_Cost'] = doctor_group['Avg_Cost'].map('₹{:,.0f}'.format)
print(doctor_group.to_string(index=False))

busiest = doctor_group.iloc[0]['doctor']
print(f"\n  ✅ Doctor treating MOST patients: {busiest}")

# ─── 5. Average Cost per Disease ────────────────────────────────
avg_cost_disease = df.groupby('disease')['treatment_cost'].mean().reset_index()
avg_cost_disease.columns = ['Disease', 'Avg Cost']

# ─── MATPLOTLIB CHARTS ──────────────────────────────────────────
colors_pie  = ['#3B82F6', '#EF4444', '#10B981', '#F59E0B']
colors_bar  = ['#6366F1', '#EC4899', '#14B8A6']
bg_color    = '#0F172A'
text_color  = '#E2E8F0'
grid_color  = '#1E293B'

# PIE CHART — Disease Distribution
fig, ax = plt.subplots(figsize=(7, 6), facecolor=bg_color)
ax.set_facecolor(bg_color)
wedges, texts, autotexts = ax.pie(
    disease_count['Patient Count'],
    labels=disease_count['Disease'],
    autopct='%1.1f%%',
    colors=colors_pie,
    startangle=140,
    wedgeprops=dict(edgecolor='#0F172A', linewidth=2),
    textprops={'color': text_color, 'fontsize': 11, 'fontweight': 'bold'}
)
for at in autotexts:
    at.set_color('#0F172A')
    at.set_fontsize(10)
    at.set_fontweight('bold')
ax.set_title('Disease Distribution', color=text_color, fontsize=14, fontweight='bold', pad=15)
plt.tight_layout()
plt.savefig('/home/claude/pie_chart.png', dpi=150, bbox_inches='tight', facecolor=bg_color)
plt.close()
print("\n✅ Pie chart saved.")

# BAR CHART — Doctor Workload
doc_counts = df['doctor'].value_counts()
fig, ax = plt.subplots(figsize=(8, 5), facecolor=bg_color)
ax.set_facecolor(bg_color)
bars = ax.bar(doc_counts.index, doc_counts.values, color=colors_bar, edgecolor='#0F172A', linewidth=1.5, width=0.5)
for bar, val in zip(bars, doc_counts.values):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.05,
            str(val), ha='center', va='bottom', color=text_color, fontsize=12, fontweight='bold')
ax.set_title('Doctor Workload (Patient Count)', color=text_color, fontsize=14, fontweight='bold', pad=15)
ax.set_xlabel('Doctor', color=text_color, fontsize=11)
ax.set_ylabel('Number of Patients', color=text_color, fontsize=11)
ax.tick_params(colors=text_color)
ax.spines['bottom'].set_color(grid_color)
ax.spines['left'].set_color(grid_color)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.set_facecolor(bg_color)
ax.yaxis.set_major_locator(plt.MaxNLocator(integer=True))
ax.grid(axis='y', color=grid_color, linewidth=0.8, alpha=0.6)
plt.tight_layout()
plt.savefig('/home/claude/bar_chart.png', dpi=150, bbox_inches='tight', facecolor=bg_color)
plt.close()
print("✅ Bar chart saved.")

# ─── EXCEL with Pivot Table ──────────────────────────────────────
wb = Workbook()

# --- Sheet 1: Raw Data ---
ws_data = wb.active
ws_data.title = "Patient Data"

header_fill = PatternFill("solid", fgColor="1E3A5F")
header_font = Font(bold=True, color="FFFFFF", name="Calibri")
border = Border(
    left=Side(style='thin', color='C0C0C0'),
    right=Side(style='thin', color='C0C0C0'),
    top=Side(style='thin', color='C0C0C0'),
    bottom=Side(style='thin', color='C0C0C0')
)

headers = list(df.columns)
for col, h in enumerate(headers, 1):
    cell = ws_data.cell(row=1, column=col, value=h.replace('_', ' ').title())
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

alt_fill = PatternFill("solid", fgColor="EBF3FC")
for r_idx, row in df.iterrows():
    for c_idx, val in enumerate(row, 1):
        cell = ws_data.cell(row=r_idx+2, column=c_idx, value=val)
        cell.font = Font(name="Calibri")
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        if r_idx % 2 == 0:
            cell.fill = alt_fill

col_widths = [12, 8, 10, 16, 14, 16]
for i, w in enumerate(col_widths, 1):
    ws_data.column_dimensions[chr(64+i)].width = w

# --- Sheet 2: Pivot — Disease Count ---
ws_pivot = wb.create_sheet("Pivot - Disease Count")
ws_pivot.cell(1, 1, "Disease").font = Font(bold=True, color="FFFFFF", name="Calibri")
ws_pivot.cell(1, 1).fill = PatternFill("solid", fgColor="0D4F8B")
ws_pivot.cell(1, 1).alignment = Alignment(horizontal='center')
ws_pivot.cell(1, 2, "Patient Count").font = Font(bold=True, color="FFFFFF", name="Calibri")
ws_pivot.cell(1, 2).fill = PatternFill("solid", fgColor="0D4F8B")
ws_pivot.cell(1, 2).alignment = Alignment(horizontal='center')

for i, row in disease_count.iterrows():
    ws_pivot.cell(i+2, 1, row['Disease']).alignment = Alignment(horizontal='center')
    ws_pivot.cell(i+2, 2, row['Patient Count']).alignment = Alignment(horizontal='center')
    ws_pivot.cell(i+2, 1).border = border
    ws_pivot.cell(i+2, 2).border = border

ws_pivot.column_dimensions['A'].width = 18
ws_pivot.column_dimensions['B'].width = 16

# Add Pie Chart to Excel
pie = PieChart()
pie.title = "Disease Distribution"
pie.style = 10
data_ref = Reference(ws_pivot, min_col=2, min_row=1, max_row=len(disease_count)+1)
cats_ref = Reference(ws_pivot, min_col=1, min_row=2, max_row=len(disease_count)+1)
pie.add_data(data_ref, titles_from_data=True)
pie.set_categories(cats_ref)
pie.width = 15
pie.height = 12
ws_pivot.add_chart(pie, "D2")

# --- Sheet 3: Pivot — Doctor Workload ---
ws_doc = wb.create_sheet("Pivot - Doctor Workload")
doc_raw = df.groupby('doctor')['patient_id'].count().reset_index()
doc_raw.columns = ['Doctor', 'Patient Count']

ws_doc.cell(1, 1, "Doctor").font = Font(bold=True, color="FFFFFF", name="Calibri")
ws_doc.cell(1, 1).fill = PatternFill("solid", fgColor="145A32")
ws_doc.cell(1, 1).alignment = Alignment(horizontal='center')
ws_doc.cell(1, 2, "Patient Count").font = Font(bold=True, color="FFFFFF", name="Calibri")
ws_doc.cell(1, 2).fill = PatternFill("solid", fgColor="145A32")
ws_doc.cell(1, 2).alignment = Alignment(horizontal='center')

for i, row in doc_raw.iterrows():
    ws_doc.cell(i+2, 1, row['Doctor']).alignment = Alignment(horizontal='center')
    ws_doc.cell(i+2, 2, row['Patient Count']).alignment = Alignment(horizontal='center')
    ws_doc.cell(i+2, 1).border = border
    ws_doc.cell(i+2, 2).border = border

ws_doc.column_dimensions['A'].width = 16
ws_doc.column_dimensions['B'].width = 16

# Add Bar Chart to Excel
bar_chart = BarChart()
bar_chart.type = "col"
bar_chart.title = "Doctor Workload"
bar_chart.y_axis.title = "Patient Count"
bar_chart.x_axis.title = "Doctor"
bar_chart.style = 10
bar_data = Reference(ws_doc, min_col=2, min_row=1, max_row=len(doc_raw)+1)
bar_cats = Reference(ws_doc, min_col=1, min_row=2, max_row=len(doc_raw)+1)
bar_chart.add_data(bar_data, titles_from_data=True)
bar_chart.set_categories(bar_cats)
bar_chart.width = 15
bar_chart.height = 12
ws_doc.add_chart(bar_chart, "D2")

# --- Sheet 4: Summary Stats ---
ws_stats = wb.create_sheet("Cost Summary")
ws_stats.cell(1, 1, "TREATMENT COST ANALYSIS").font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
ws_stats.cell(1, 1).fill = PatternFill("solid", fgColor="7B2D8B")
ws_stats.cell(1, 1).alignment = Alignment(horizontal='center')
ws_stats.merge_cells('A1:B1')

summary = [
    ("Average Cost", f"₹{df['treatment_cost'].mean():,.0f}"),
    ("Minimum Cost", f"₹{df['treatment_cost'].min():,.0f}"),
    ("Maximum Cost", f"₹{df['treatment_cost'].max():,.0f}"),
    ("Median Cost",  f"₹{df['treatment_cost'].median():,.0f}"),
    ("Total Revenue",f"₹{df['treatment_cost'].sum():,.0f}"),
]
for i, (label, val) in enumerate(summary, 2):
    ws_stats.cell(i, 1, label).font = Font(bold=True, name="Calibri")
    ws_stats.cell(i, 1).border = border
    ws_stats.cell(i, 1).fill = PatternFill("solid", fgColor="F3E5F5")
    ws_stats.cell(i, 2, val).alignment = Alignment(horizontal='center')
    ws_stats.cell(i, 2).border = border
    ws_stats.cell(i, 2).font = Font(name="Calibri")

ws_stats.column_dimensions['A'].width = 18
ws_stats.column_dimensions['B'].width = 16

wb.save('/home/claude/Hospital_Analysis.xlsx')
print("✅ Excel workbook saved with pivot tables and charts.")
print("\n" + "=" * 55)
print("  ANALYSIS COMPLETE")
print("=" * 55)
